"""
Excel database utilities for booking app
Handles all database operations with Excel files
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from typing import Dict, List, Optional
from config import DB_FILE, FLIGHT_FIELDS, HOTEL_FIELDS, TRAIN_FIELDS, BUS_FIELDS


class ExcelDatabase:
    """Handle all Excel database operations"""

    def __init__(self, db_path: str = str(DB_FILE)):
        self.db_path = db_path
        self.ensure_database_exists()

    def ensure_database_exists(self):
        """Create database file with sheets if it doesn't exist"""
        if not os.path.exists(self.db_path):
            wb = Workbook()
            if wb.active is not None:
                wb.remove(wb.active)  # Remove default sheet

            # Create sheets for each booking type
            sheets = {
                "Flight": FLIGHT_FIELDS,
                "Hotel": HOTEL_FIELDS,
                "Train": TRAIN_FIELDS,
                "Bus": BUS_FIELDS,
            }

            for sheet_name, fields in sheets.items():
                ws = wb.create_sheet(sheet_name)
                self._format_header(ws, fields)

            wb.save(self.db_path)

    def _format_header(self, worksheet, headers: List[str]):
        """Format header row with styling"""
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Set column widths
        for col_num in range(1, len(headers) + 1):
            worksheet.column_dimensions[
                worksheet.cell(1, col_num).column_letter
            ].width = 18

    def add_booking(self, booking_type: str, data: Dict) -> tuple[bool, str]:
        """
        Add a new booking record
        Returns: (success: bool, message: str)
        """
        try:
            sheet_name = booking_type.capitalize()
            df = pd.read_excel(self.db_path, sheet_name=sheet_name)

            # Add new row
            new_row = pd.DataFrame([data])
            df = pd.concat([df, new_row], ignore_index=True)

            # Write back to Excel
            with pd.ExcelWriter(
                self.db_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Reformat headers
            wb = load_workbook(self.db_path)
            ws = wb[sheet_name]
            fields = self._get_fields_for_type(booking_type)
            self._format_header(ws, fields)
            wb.save(self.db_path)

            return True, f"✅ {booking_type.capitalize()} booking added successfully!"

        except Exception as e:
            return False, f"❌ Error adding booking: {str(e)}"

    def get_bookings(self, booking_type: str) -> pd.DataFrame:
        """Get all bookings of a specific type"""
        try:
            sheet_name = booking_type.capitalize()
            df = pd.read_excel(self.db_path, sheet_name=sheet_name)
            return df
        except Exception as e:
            print(f"Error reading bookings: {e}")
            return pd.DataFrame()

    def update_booking(
        self, booking_type: str, booking_id: str, data: Dict
    ) -> tuple[bool, str]:
        """
        Update an existing booking record
        Returns: (success: bool, message: str)
        """
        try:
            sheet_name = booking_type.capitalize()
            df = pd.read_excel(self.db_path, sheet_name=sheet_name)

            # Find and update the booking
            if "Booking ID" not in df.columns:
                return False, "Booking ID column not found"

            mask = df["Booking ID"] == booking_id
            if not mask.any():
                return False, f"Booking ID {booking_id} not found"

            # Update the row
            index = df[mask].index[0]
            for key, value in data.items():
                if key in df.columns:
                    df.at[index, key] = value

            # Write back to Excel
            with pd.ExcelWriter(
                self.db_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Reformat headers
            wb = load_workbook(self.db_path)
            ws = wb[sheet_name]
            fields = self._get_fields_for_type(booking_type)
            self._format_header(ws, fields)
            wb.save(self.db_path)

            return True, f"✅ {booking_type.capitalize()} booking updated successfully!"

        except Exception as e:
            return False, f"❌ Error updating booking: {str(e)}"

    def delete_booking(self, booking_type: str, booking_id: str) -> tuple[bool, str]:
        """
        Delete a booking record
        Returns: (success: bool, message: str)
        """
        try:
            sheet_name = booking_type.capitalize()
            df = pd.read_excel(self.db_path, sheet_name=sheet_name)

            if "Booking ID" not in df.columns:
                return False, "Booking ID column not found"

            initial_len = len(df)
            df = df[df["Booking ID"] != booking_id]

            if len(df) == initial_len:
                return False, f"Booking ID {booking_id} not found"

            # Write back to Excel
            with pd.ExcelWriter(
                self.db_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Reformat headers
            wb = load_workbook(self.db_path)
            ws = wb[sheet_name]
            fields = self._get_fields_for_type(booking_type)
            self._format_header(ws, fields)
            wb.save(self.db_path)

            return True, f"✅ {booking_type.capitalize()} booking deleted successfully!"

        except Exception as e:
            return False, f"❌ Error deleting booking: {str(e)}"

    def search_bookings(
        self, booking_type: str, search_field: str, search_value: str
    ) -> pd.DataFrame:
        """Search bookings by a specific field"""
        try:
            df = self.get_bookings(booking_type)

            if search_field not in df.columns:
                return pd.DataFrame()

            # Case-insensitive search
            mask = (
                df[search_field]
                .astype(str)
                .str.contains(search_value, case=False, na=False)
            )
            return df[mask]

        except Exception as e:
            print(f"Error searching bookings: {e}")
            return pd.DataFrame()

    def filter_by_date_range(
        self, booking_type: str, date_field: str, start_date, end_date
    ) -> pd.DataFrame:
        """Filter bookings by date range"""
        try:
            df = self.get_bookings(booking_type)

            if date_field not in df.columns:
                return pd.DataFrame()

            df[date_field] = pd.to_datetime(df[date_field], errors="coerce")
            mask = (df[date_field] >= start_date) & (df[date_field] <= end_date)
            return df[mask]

        except Exception as e:
            print(f"Error filtering by date: {e}")
            return pd.DataFrame()

    def get_statistics(self, booking_type: str) -> Dict:
        """Get statistics for a booking type"""
        try:
            df = self.get_bookings(booking_type)

            stats = {
                "total_bookings": len(df),
                "confirmed": (
                    len(df[df["Status"] == "Confirmed"])
                    if "Status" in df.columns
                    else 0
                ),
                "pending": (
                    len(df[df["Status"] == "Pending"]) if "Status" in df.columns else 0
                ),
                "cancelled": (
                    len(df[df["Status"] == "Cancelled"])
                    if "Status" in df.columns
                    else 0
                ),
            }

            if "Total Cost" in df.columns:
                stats["total_revenue"] = df["Total Cost"].sum()

            return stats

        except Exception as e:
            print(f"Error getting statistics: {e}")
            return {}

    def export_to_csv(self, booking_type: str, output_path: str) -> tuple[bool, str]:
        """Export bookings to CSV"""
        try:
            df = self.get_bookings(booking_type)
            df.to_csv(output_path, index=False)
            return True, f"✅ Exported to {output_path}"
        except Exception as e:
            return False, f"❌ Error exporting: {str(e)}"

    @staticmethod
    def _get_fields_for_type(booking_type: str) -> List[str]:
        """Get field list for booking type"""
        type_map = {
            "flight": FLIGHT_FIELDS,
            "hotel": HOTEL_FIELDS,
            "train": TRAIN_FIELDS,
            "bus": BUS_FIELDS,
        }
        return type_map.get(booking_type.lower(), [])

    def generate_booking_id(self, booking_type: str) -> str:
        """Generate unique booking ID"""
        prefix = booking_type[:2].upper()
        df = self.get_bookings(booking_type)

        if len(df) == 0:
            return f"{prefix}001"

        try:
            last_id = df["Booking ID"].iloc[-1]
            if isinstance(last_id, str) and last_id.startswith(prefix):
                number = int(last_id[2:]) + 1
                return f"{prefix}{number:03d}"
        except (ValueError, IndexError):
            pass

        return f"{prefix}{len(df) + 1:03d}"
